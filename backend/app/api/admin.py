"""
Endpoints pour gestion Professeurs, Classes, Matières (Admin établissement)
"""
import csv
import io
import re

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr, ConfigDict, Field
from typing import List, Optional
from datetime import datetime

from app.db.connection import get_db_session
from app.db.multi_tenant import get_tenant_session
from app.auth.security import get_current_user, hash_password

router = APIRouter(prefix="/admin", tags=["Admin Management"])


# ════════════════════════════════════════════════════════════
# Schémas Pydantic
# ════════════════════════════════════════════════════════════

class ProfesseurCreate(BaseModel):
    nom: str
    prenom: str
    email: EmailStr
    phone: Optional[str] = None
    specialite: Optional[str] = None
    matricule: str
    username: Optional[str] = None
    password: Optional[str] = None
    section: Optional[str] = "francophone"


class ProfesseurUpdate(BaseModel):
    nom: Optional[str] = None
    prenom: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    specialite: Optional[str] = None
    section: Optional[str] = None


class ProfesseurResponse(BaseModel):
    id: int
    nom: str
    prenom: str
    email: str
    phone: Optional[str] = None
    specialite: Optional[str] = None
    matricule: str
    section: Optional[str] = "francophone"
    is_active: bool
    created_at: datetime = Field(validation_alias="date_creation")

    model_config = ConfigDict(from_attributes=True, populate_by_name=True)


class AttributionProfCreate(BaseModel):
    professeur_id: int
    classe_id: int
    matiere_id: int


class AttributionProfResponse(BaseModel):
    id: int
    professeur_id: int
    classe_id: int
    matiere_id: int
    is_active: bool

    model_config = ConfigDict(from_attributes=True)


class AnneeScolaireCreate(BaseModel):
    annee: str
    date_debut: datetime
    date_fin: datetime
    is_active: Optional[bool] = True


class AnneeScolaireResponse(BaseModel):
    id: int
    annee: str
    date_debut: datetime
    date_fin: datetime
    is_active: bool

    model_config = ConfigDict(from_attributes=True)


class EleveCreate(BaseModel):
    nom: str
    prenom: str
    matricule: str
    classe_id: Optional[int] = None
    section: Optional[str] = None


class EleveUpdate(BaseModel):
    nom: Optional[str] = None
    prenom: Optional[str] = None
    classe_id: Optional[int] = None
    section: Optional[str] = None


class EleveResponse(BaseModel):
    id: int
    nom: str
    prenom: str
    matricule: str
    classe_id: Optional[int] = None
    date_inscription: datetime
    classe_nom: Optional[str] = None
    section: Optional[str] = None

    model_config = ConfigDict(from_attributes=True)


class EleveImportResult(BaseModel):
    created: int
    updated: int
    total: int
    errors: List[str]


ELEVE_HEADER_ALIASES = {
    "nom": ["nom", "name", "lastname", "last name", "nom de famille", "surname"],
    "prenom": ["prenom", "prénom", "firstname", "first name", "given name"],
    "matricule": ["matricule", "id", "unique id", "unique_id", "matricule unique", "code"],
    "classe": ["classe", "class"],
    "section": ["section"],
    "sexe": ["sexe", "sex", "gender"],
    "redoublant": ["redoublant", "repeater", "redouble"],
}


def _normalize_header_cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _parse_eleve_header_row(row) -> dict:
    mapping = {}
    for idx, cell in enumerate(row):
        key = _normalize_header_cell(cell)
        if not key:
            continue
        for field, aliases in ELEVE_HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _parse_repeater(value) -> bool:
    if value is None or str(value).strip() == "":
        return False
    return str(value).strip().upper() in ("OUI", "YES", "Y", "TRUE", "1", "O")


def _parse_sexe(value) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip().upper()
    if s in ("M", "H", "MALE", "GARCON", "BOY"):
        return "M"
    if s in ("F", "FEMALE", "FILLE", "GIRL"):
        return "F"
    return None


def _read_eleve_import_rows(content: bytes, filename: str) -> list:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return [list(row) for row in csv.reader(io.StringIO(text))]
    if lower.endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=501,
                detail="Import Excel indisponible (openpyxl non installé). Utilisez un fichier CSV.",
            ) from exc
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Fichier Excel invalide : {exc}") from exc
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    raise HTTPException(status_code=400, detail="Fichier Excel (.xlsx) ou CSV (.csv) requis")


class ClasseCreate(BaseModel):
    nom: str
    niveau: str
    annee_scolaire_id: Optional[int] = None
    capacite: Optional[int] = 30
    salle: Optional[str] = None
    section: Optional[str] = "francophone"
    serie: Optional[str] = None


class ClasseUpdate(BaseModel):
    nom: Optional[str] = None
    niveau: Optional[str] = None
    capacite: Optional[int] = None
    salle: Optional[str] = None
    section: Optional[str] = None
    serie: Optional[str] = None


class ClasseResponse(BaseModel):
    id: int
    nom: str
    niveau: str
    capacite: int
    salle: Optional[str] = None
    section: Optional[str] = "francophone"
    serie: Optional[str] = None
    created_at: datetime

    model_config = ConfigDict(from_attributes=True)


class MatiereCreate(BaseModel):
    nom: str
    code: str
    description: Optional[str] = None
    groupe: Optional[int] = 1
    coefficient_defaut: Optional[float] = 1.0


class MatiereUpdate(BaseModel):
    nom: Optional[str] = None
    code: Optional[str] = None
    description: Optional[str] = None
    groupe: Optional[int] = None
    coefficient_defaut: Optional[float] = None


class MatiereResponse(BaseModel):
    id: int
    nom: str
    code: str
    description: Optional[str] = None
    groupe: Optional[int] = 1
    coefficient_defaut: Optional[float] = 1.0

    model_config = ConfigDict(from_attributes=True)


class BulletinSettingsResponse(BaseModel):
    school_id: int
    name: str
    logo_url: Optional[str] = None
    bulletin_po_box: Optional[str] = None
    bulletin_motto: Optional[str] = None
    bulletin_delegation_en: Optional[str] = None
    bulletin_delegation_fr: Optional[str] = None
    bulletin_next_term_note: Optional[str] = None
    bulletin_template: str = "cameroon_bilingual"
    bulletin_scope: str = "trimestre"
    available_templates: dict
    available_scopes: dict


class BulletinSettingsUpdate(BaseModel):
    logo_url: Optional[str] = None
    bulletin_po_box: Optional[str] = None
    bulletin_motto: Optional[str] = None
    bulletin_delegation_en: Optional[str] = None
    bulletin_delegation_fr: Optional[str] = None
    bulletin_next_term_note: Optional[str] = None
    bulletin_template: Optional[str] = None
    bulletin_scope: Optional[str] = None


# ════════════════════════════════════════════════════════════
# PROFESSEURS — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/professeurs/", response_model=ProfesseurResponse, status_code=status.HTTP_201_CREATED)
def create_professeur(
    prof_data: ProfesseurCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Crée un nouveau professeur"""
    from app.models.school import Professeur
    
    # Vérifier que l'utilisateur est admin
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Seul un administrateur peut créer des professeurs"
        )
    
    # Vérifier si le matricule existe
    existing = db.query(Professeur).filter(Professeur.matricule == prof_data.matricule).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ce matricule existe déjà"
        )
    
    # Vérifier si l'email existe
    existing_email = db.query(Professeur).filter(Professeur.email == prof_data.email).first()
    if existing_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cet email est déjà utilisé"
        )
    
    from app.services.sections import validate_section

    try:
        try:
            prof_section = validate_section(prof_data.section, allow_both=True)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
        professeur = Professeur(
            nom=prof_data.nom,
            prenom=prof_data.prenom,
            email=prof_data.email,
            phone=prof_data.phone,
            specialite=prof_data.specialite,
            matricule=prof_data.matricule,
            section=prof_section,
            username=prof_data.username or prof_data.email.split('@')[0],
            hashed_password=hash_password(prof_data.password or "default123"),
            is_active=True
        )
        db.add(professeur)
        db.commit()
        db.refresh(professeur)
        return professeur
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur création professeur: {str(e)}"
        )


@router.get("/professeurs/", response_model=List[ProfesseurResponse])
def list_professeurs(
    current_user: dict = Depends(get_current_user),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_tenant_session)
):
    """Liste tous les professeurs"""
    from app.models.school import Professeur
    
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non autorisé"
        )
    
    professeurs = db.query(Professeur).offset(skip).limit(limit).all()
    return professeurs


@router.get("/professeurs/{prof_id}", response_model=ProfesseurResponse)
def get_professeur(
    prof_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Récupère les détails d'un professeur"""
    from app.models.school import Professeur
    
    professeur = db.query(Professeur).filter(Professeur.id == prof_id).first()
    if not professeur:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Professeur non trouvé"
        )
    
    return professeur


@router.put("/professeurs/{prof_id}", response_model=ProfesseurResponse)
def update_professeur(
    prof_id: int,
    prof_update: ProfesseurUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Modifie un professeur"""
    from app.models.school import Professeur
    
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non autorisé"
        )
    
    professeur = db.query(Professeur).filter(Professeur.id == prof_id).first()
    if not professeur:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Professeur non trouvé"
        )
    
    from app.services.sections import validate_section

    update_data = prof_update.dict(exclude_unset=True)
    if "section" in update_data:
        try:
            update_data["section"] = validate_section(update_data["section"], allow_both=True)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    for field, value in update_data.items():
        setattr(professeur, field, value)
    
    db.commit()
    db.refresh(professeur)
    return professeur


def _purge_professeur_dependencies(db, professeur_id: int) -> None:
    from app.models.school import Note, AttributionProfesseur, EmploiTemps
    db.query(Note).filter(Note.professeur_id == professeur_id).delete(synchronize_session=False)
    db.query(AttributionProfesseur).filter(
        AttributionProfesseur.professeur_id == professeur_id,
    ).delete(synchronize_session=False)
    db.query(EmploiTemps).filter(EmploiTemps.professeur_id == professeur_id).delete(synchronize_session=False)


def _purge_matiere_dependencies(db, matiere_id: int) -> None:
    from app.models.school import Note, AttributionProfesseur, EmploiTemps, PeriodeSaisieNotes
    db.query(Note).filter(Note.matiere_id == matiere_id).delete(synchronize_session=False)
    db.query(AttributionProfesseur).filter(
        AttributionProfesseur.matiere_id == matiere_id,
    ).delete(synchronize_session=False)
    db.query(PeriodeSaisieNotes).filter(
        PeriodeSaisieNotes.matiere_id == matiere_id,
    ).delete(synchronize_session=False)
    db.query(EmploiTemps).filter(EmploiTemps.matiere_id == matiere_id).delete(synchronize_session=False)


def _purge_classe_dependencies(db, classe_id: int) -> None:
    from app.models.school import (
        Eleve, Note, Bulletin, AttributionProfesseur, EmploiTemps, PeriodeSaisieNotes,
    )
    eleve_ids = [
        row[0] for row in db.query(Eleve.id).filter(Eleve.classe_id == classe_id).all()
    ]
    if eleve_ids:
        db.query(Note).filter(Note.eleve_id.in_(eleve_ids)).delete(synchronize_session=False)
        db.query(Bulletin).filter(Bulletin.eleve_id.in_(eleve_ids)).delete(synchronize_session=False)
    db.query(Bulletin).filter(Bulletin.classe_id == classe_id).delete(synchronize_session=False)
    db.query(AttributionProfesseur).filter(
        AttributionProfesseur.classe_id == classe_id,
    ).delete(synchronize_session=False)
    db.query(PeriodeSaisieNotes).filter(
        PeriodeSaisieNotes.classe_id == classe_id,
    ).delete(synchronize_session=False)
    db.query(EmploiTemps).filter(EmploiTemps.classe_id == classe_id).delete(synchronize_session=False)
    db.query(Eleve).filter(Eleve.classe_id == classe_id).delete(synchronize_session=False)


def _purge_eleve_dependencies(db, eleve_id: int) -> None:
    from app.models.school import Note, Bulletin
    db.query(Note).filter(Note.eleve_id == eleve_id).delete(synchronize_session=False)
    db.query(Bulletin).filter(Bulletin.eleve_id == eleve_id).delete(synchronize_session=False)


@router.delete("/professeurs/{prof_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_professeur(
    prof_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Supprime un professeur et ses données liées"""
    from app.models.school import Professeur
    
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non autorisé"
        )
    
    professeur = db.query(Professeur).filter(Professeur.id == prof_id).first()
    if not professeur:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Professeur non trouvé"
        )
    
    try:
        _purge_professeur_dependencies(db, prof_id)
        db.delete(professeur)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Impossible de supprimer le professeur : {e}") from e


# ════════════════════════════════════════════════════════════
# ATTRIBUTIONS PROFESSEURS — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/attributions-professeurs/", response_model=AttributionProfResponse, status_code=status.HTTP_201_CREATED)
def create_attribution(
    attrib_data: AttributionProfCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Attribue un professeur à une classe/matière"""
    from app.models.school import AttributionProfesseur, Professeur, Classe, Matiere
    
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non autorisé"
        )
    
    # Vérifier existence
    prof = db.query(Professeur).filter(Professeur.id == attrib_data.professeur_id).first()
    if not prof:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Professeur non trouvé")
    
    classe = db.query(Classe).filter(Classe.id == attrib_data.classe_id).first()
    if not classe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Classe non trouvée")
    
    matiere = db.query(Matiere).filter(Matiere.id == attrib_data.matiere_id).first()
    if not matiere:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Matière non trouvée")

    from app.services.sections import sections_compatible
    if not sections_compatible(getattr(prof, "section", None), getattr(classe, "section", None)):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Le professeur ({getattr(prof, 'section', 'francophone')}) "
                f"ne peut pas enseigner en section {getattr(classe, 'section', 'francophone')}"
            ),
        )
    
    # Vérifier si attribution existe déjà
    existing = db.query(AttributionProfesseur).filter(
        AttributionProfesseur.professeur_id == attrib_data.professeur_id,
        AttributionProfesseur.classe_id == attrib_data.classe_id,
        AttributionProfesseur.matiere_id == attrib_data.matiere_id
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cette attribution existe déjà"
        )
    
    try:
        attribution = AttributionProfesseur(**attrib_data.dict())
        db.add(attribution)
        db.commit()
        db.refresh(attribution)
        return attribution
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur: {str(e)}"
        )


@router.get("/attributions-professeurs/", response_model=List[AttributionProfResponse])
def list_attributions(
    current_user: dict = Depends(get_current_user),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_tenant_session)
):
    """Liste les attributions"""
    from app.models.school import AttributionProfesseur
    
    attributions = db.query(AttributionProfesseur).offset(skip).limit(limit).all()
    return attributions


@router.delete("/attributions-professeurs/{attrib_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_attribution(
    attrib_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Supprime une attribution"""
    from app.models.school import AttributionProfesseur
    
    if current_user.get("role") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Non autorisé"
        )
    
    attribution = db.query(AttributionProfesseur).filter(AttributionProfesseur.id == attrib_id).first()
    if not attribution:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attribution non trouvée"
        )
    
    db.delete(attribution)
    db.commit()


# ════════════════════════════════════════════════════════════
# CLASSES — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/classes/", response_model=ClasseResponse, status_code=status.HTTP_201_CREATED)
def create_classe(
    classe_data: ClasseCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Crée une nouvelle classe"""
    from app.models.school import Classe, AnneeScolaire
    
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    
    # Si annee_scolaire_id non fourni, utiliser l'année active ou la première disponible
    annee_scolaire_id = classe_data.annee_scolaire_id
    if not annee_scolaire_id:
        annee = db.query(AnneeScolaire).filter(AnneeScolaire.is_active == True).first()
        if not annee:
            annee = db.query(AnneeScolaire).first()
        if annee:
            annee_scolaire_id = annee.id
    
    try:
        data = classe_data.dict()
        data["annee_scolaire_id"] = annee_scolaire_id
        classe = Classe(**data)
        db.add(classe)
        db.commit()
        db.refresh(classe)
        return classe
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.get("/classes/", response_model=List[ClasseResponse])
def list_classes(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Liste les classes"""
    from app.models.school import Classe
    
    classes = db.query(Classe).all()
    return classes


@router.patch("/classes/{classe_id}", response_model=ClasseResponse)
def update_classe(
    classe_id: int,
    data: ClasseUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    from app.models.school import Classe

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    classe = db.query(Classe).filter(Classe.id == classe_id).first()
    if not classe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Classe non trouvée")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(classe, field, value)
    db.commit()
    db.refresh(classe)
    return classe


@router.delete("/classes/{classe_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_classe(
    classe_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Supprime une classe"""
    from app.models.school import Classe
    
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    
    classe = db.query(Classe).filter(Classe.id == classe_id).first()
    if not classe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Classe non trouvée")
    
    try:
        _purge_classe_dependencies(db, classe_id)
        db.delete(classe)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Impossible de supprimer la classe : {e}") from e


# ════════════════════════════════════════════════════════════
# MATIÈRES — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/matieres/", response_model=MatiereResponse, status_code=status.HTTP_201_CREATED)
def create_matiere(
    matiere_data: MatiereCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Crée une nouvelle matière"""
    from app.models.school import Matiere
    
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    
    # Vérifier si code existe
    existing = db.query(Matiere).filter(Matiere.code == matiere_data.code).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ce code existe déjà")

    if matiere_data.groupe is not None and matiere_data.groupe not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="Le groupe doit être 1, 2 ou 3")
    
    try:
        matiere = Matiere(**matiere_data.dict())
        db.add(matiere)
        db.commit()
        db.refresh(matiere)
        return matiere
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.get("/matieres/", response_model=List[MatiereResponse])
def list_matieres(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Liste les matières"""
    from app.models.school import Matiere
    
    matieres = db.query(Matiere).all()
    return matieres


@router.patch("/matieres/{matiere_id}", response_model=MatiereResponse)
def update_matiere(
    matiere_id: int,
    data: MatiereUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    from app.models.school import Matiere

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    matiere = db.query(Matiere).filter(Matiere.id == matiere_id).first()
    if not matiere:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Matière non trouvée")

    payload = data.model_dump(exclude_unset=True)
    if "groupe" in payload and payload["groupe"] not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="Le groupe doit être 1, 2 ou 3")
    if "code" in payload:
        existing = db.query(Matiere).filter(
            Matiere.code == payload["code"], Matiere.id != matiere_id,
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Ce code existe déjà")

    for field, value in payload.items():
        setattr(matiere, field, value)
    db.commit()
    db.refresh(matiere)
    return matiere


@router.delete("/matieres/{matiere_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_matiere(
    matiere_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session)
):
    """Supprime une matière"""
    from app.models.school import Matiere
    
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    
    matiere = db.query(Matiere).filter(Matiere.id == matiere_id).first()
    if not matiere:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Matière non trouvée")
    
    try:
        _purge_matiere_dependencies(db, matiere_id)
        db.delete(matiere)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Impossible de supprimer la matière : {e}") from e


# ════════════════════════════════════════════════════════════
# ANNÉES SCOLAIRES — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/annees-scolaires/", response_model=AnneeScolaireResponse, status_code=status.HTTP_201_CREATED)
def create_annee_scolaire(
    data: AnneeScolaireCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Crée une année scolaire"""
    from app.models.school import AnneeScolaire

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    # Si is_active=True, désactiver les autres
    if data.is_active:
        db.query(AnneeScolaire).filter(AnneeScolaire.is_active == True).update({"is_active": False})

    annee = AnneeScolaire(**data.dict())
    db.add(annee)
    db.commit()
    db.refresh(annee)
    return annee


@router.get("/annees-scolaires/", response_model=List[AnneeScolaireResponse])
def list_annees_scolaires(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Liste les années scolaires"""
    from app.models.school import AnneeScolaire

    return db.query(AnneeScolaire).order_by(AnneeScolaire.date_debut.desc()).all()


@router.put("/annees-scolaires/{annee_id}/activer", response_model=AnneeScolaireResponse)
def activer_annee_scolaire(
    annee_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Définit une année scolaire comme active"""
    from app.models.school import AnneeScolaire

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    db.query(AnneeScolaire).filter(AnneeScolaire.is_active == True).update({"is_active": False})
    annee = db.query(AnneeScolaire).filter(AnneeScolaire.id == annee_id).first()
    if not annee:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Année non trouvée")
    annee.is_active = True
    db.commit()
    db.refresh(annee)
    return annee


@router.delete("/annees-scolaires/{annee_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_annee_scolaire(
    annee_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Supprime une année scolaire"""
    from app.models.school import AnneeScolaire

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    annee = db.query(AnneeScolaire).filter(AnneeScolaire.id == annee_id).first()
    if not annee:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Année non trouvée")

    db.delete(annee)
    db.commit()


def _eleve_to_response(eleve, classe=None) -> EleveResponse:
    classe_nom = None
    section = None
    if classe:
        classe_nom = f"{classe.nom} ({classe.niveau})"
        section = getattr(classe, "section", None) or "francophone"
    return EleveResponse(
        id=eleve.id,
        nom=eleve.nom,
        prenom=eleve.prenom,
        matricule=eleve.matricule,
        classe_id=eleve.classe_id,
        date_inscription=eleve.date_inscription,
        classe_nom=classe_nom,
        section=section,
    )


# ════════════════════════════════════════════════════════════
# ÉLÈVES — CRUD
# ════════════════════════════════════════════════════════════

@router.post("/eleves/", response_model=EleveResponse, status_code=status.HTTP_201_CREATED)
def create_eleve(
    eleve_data: EleveCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Crée un nouvel élève"""
    from app.models.school import Eleve, Classe

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    existing = db.query(Eleve).filter(Eleve.matricule == eleve_data.matricule).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ce matricule existe déjà")

    classe = None
    if eleve_data.classe_id:
        classe = db.query(Classe).filter(Classe.id == eleve_data.classe_id).first()
        if not classe:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Classe non trouvée")
        if eleve_data.section:
            classe_section = getattr(classe, "section", None) or "francophone"
            if eleve_data.section != classe_section:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"La classe {classe.nom} est en section {classe_section}, pas {eleve_data.section}",
                )

    try:
        payload = eleve_data.model_dump(exclude={"section"})
        eleve = Eleve(**payload)
        db.add(eleve)
        db.commit()
        db.refresh(eleve)

        if not classe and eleve.classe_id:
            classe = db.query(Classe).filter(Classe.id == eleve.classe_id).first()

        return _eleve_to_response(eleve, classe)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.get("/eleves/", response_model=List[EleveResponse])
def list_eleves(
    current_user: dict = Depends(get_current_user),
    classe_id: Optional[int] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_tenant_session),
):
    """Liste les élèves avec filtre optionnel par classe et recherche"""
    from app.models.school import Eleve, Classe

    query = db.query(Eleve)

    if classe_id:
        query = query.filter(Eleve.classe_id == classe_id)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (Eleve.nom.ilike(pattern))
            | (Eleve.prenom.ilike(pattern))
            | (Eleve.matricule.ilike(pattern))
        )

    eleves = query.offset(skip).limit(limit).all()

    result = []
    for e in eleves:
        classe = db.query(Classe).filter(Classe.id == e.classe_id).first() if e.classe_id else None
        result.append(_eleve_to_response(e, classe))
    return result


@router.get("/eleves/{eleve_id}", response_model=EleveResponse)
def get_eleve(
    eleve_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Récupère un élève"""
    from app.models.school import Eleve, Classe

    eleve = db.query(Eleve).filter(Eleve.id == eleve_id).first()
    if not eleve:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Élève non trouvé")

    classe = db.query(Classe).filter(Classe.id == eleve.classe_id).first() if eleve.classe_id else None
    return _eleve_to_response(eleve, classe)


@router.put("/eleves/{eleve_id}", response_model=EleveResponse)
def update_eleve(
    eleve_id: int,
    eleve_update: EleveUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Modifie un élève (nom, prénom, classe)"""
    from app.models.school import Eleve, Classe

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    eleve = db.query(Eleve).filter(Eleve.id == eleve_id).first()
    if not eleve:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Élève non trouvé")

    classe = None
    if eleve_update.classe_id is not None and eleve_update.classe_id != 0:
        classe = db.query(Classe).filter(Classe.id == eleve_update.classe_id).first()
        if not classe:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Classe non trouvée")
        if eleve_update.section:
            classe_section = getattr(classe, "section", None) or "francophone"
            if eleve_update.section != classe_section:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"La classe {classe.nom} est en section {classe_section}",
                )

    for field, value in eleve_update.dict(exclude_unset=True, exclude={"section"}).items():
        setattr(eleve, field, value)

    db.commit()
    db.refresh(eleve)

    if not classe and eleve.classe_id:
        classe = db.query(Classe).filter(Classe.id == eleve.classe_id).first()

    return _eleve_to_response(eleve, classe)


@router.delete("/eleves/{eleve_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_eleve(
    eleve_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Supprime un élève"""
    from app.models.school import Eleve

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    eleve = db.query(Eleve).filter(Eleve.id == eleve_id).first()
    if not eleve:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Élève non trouvé")

    try:
        _purge_eleve_dependencies(db, eleve_id)
        db.delete(eleve)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Impossible de supprimer l'élève : {e}") from e


@router.get("/eleves/import/template.xlsx")
def download_eleves_import_template(
    current_user: dict = Depends(get_current_user),
):
    """Télécharge un modèle Excel pour l'import groupé des élèves."""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(
            status_code=501,
            detail="Modèle Excel indisponible (openpyxl non installé).",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Eleves"
    headers = ["Matricule", "Nom", "Prénom", "Classe", "Section", "Sexe", "Redoublant"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["DT12345", "DOUANLA", "Linda", "Form 1A", "anglophone", "F", "NON"])
    ws.append(["DT12346", "KAMGA", "Paul", "Form 1A", "anglophone", "M", "NON"])
    ws.append([])
    ws.append(["Section : francophone ou anglophone"])
    ws.append(["Sexe : M ou F — Redoublant : OUI ou NON"])
    ws.append(["La colonne Classe est optionnelle si vous importez pour une classe déjà choisie"])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 36)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modele_import_eleves.xlsx"'},
    )


@router.post("/eleves/import", response_model=EleveImportResult)
async def import_eleves(
    file: UploadFile = File(...),
    default_classe_id: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Importe une liste d'élèves depuis Excel ou CSV."""
    from app.models.school import Eleve, Classe

    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")

    if not file.filename:
        raise HTTPException(status_code=400, detail="Fichier requis")

    default_classe = None
    if default_classe_id:
        default_classe = db.query(Classe).filter(Classe.id == default_classe_id).first()
        if not default_classe:
            raise HTTPException(status_code=404, detail="Classe par défaut introuvable")

    content = await file.read()
    rows = _read_eleve_import_rows(content, file.filename)
    if not rows:
        raise HTTPException(status_code=400, detail="Fichier vide")

    header_idx = None
    col_map = {}
    for i, row in enumerate(rows[:15]):
        candidate = _parse_eleve_header_row(list(row))
        if "matricule" in candidate and "nom" in candidate and "prenom" in candidate:
            header_idx = i
            col_map = candidate
            break

    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="En-têtes requis : Matricule, Nom, Prénom (Classe, Section, Sexe, Redoublant optionnels)",
        )

    classes_by_name = {
        c.nom.strip().lower(): c
        for c in db.query(Classe).all()
    }
    existing_by_matricule = {
        e.matricule.strip().upper(): e
        for e in db.query(Eleve).all()
    }

    created = 0
    updated = 0
    errors: List[str] = []

    for line_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            matricule = str(row[col_map["matricule"]]).strip().upper()
            nom = str(row[col_map["nom"]]).strip()
            prenom = str(row[col_map["prenom"]]).strip()
            if not matricule or not nom or not prenom:
                errors.append(f"Ligne {line_no} : matricule, nom et prénom obligatoires")
                continue

            classe = default_classe
            section_hint = None
            if "section" in col_map and row[col_map["section"]] is not None:
                section_hint = str(row[col_map["section"]]).strip().lower()
                if section_hint in ("en", "anglais", "english"):
                    section_hint = "anglophone"
                elif section_hint in ("fr", "français", "francais", "french"):
                    section_hint = "francophone"

            if "classe" in col_map and row[col_map["classe"]] is not None:
                classe_name = str(row[col_map["classe"]]).strip().lower()
                if classe_name:
                    classe = classes_by_name.get(classe_name)
                    if not classe:
                        errors.append(f"Ligne {line_no} : classe « {row[col_map['classe']]} » introuvable")
                        continue

            sexe = _parse_sexe(row[col_map["sexe"]]) if "sexe" in col_map else None
            redoublant = _parse_repeater(row[col_map["redoublant"]]) if "redoublant" in col_map else False

            if classe and section_hint:
                classe_section = getattr(classe, "section", None) or "francophone"
                if section_hint != classe_section:
                    errors.append(
                        f"Ligne {line_no} : section {section_hint} incompatible avec la classe {classe.nom}"
                    )
                    continue

            existing = existing_by_matricule.get(matricule)
            if existing:
                existing.nom = nom
                existing.prenom = prenom
                if classe:
                    existing.classe_id = classe.id
                if sexe:
                    existing.sexe = sexe
                existing.redoublant = redoublant
                updated += 1
            else:
                eleve = Eleve(
                    nom=nom,
                    prenom=prenom,
                    matricule=matricule,
                    classe_id=classe.id if classe else None,
                    sexe=sexe,
                    redoublant=redoublant,
                )
                db.add(eleve)
                existing_by_matricule[matricule] = eleve
                created += 1
        except Exception as exc:
            errors.append(f"Ligne {line_no} : {exc}")

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'enregistrement : {exc}") from exc

    return EleveImportResult(
        created=created,
        updated=updated,
        total=created + updated,
        errors=errors[:50],
    )


# ════════════════════════════════════════════════════════════
# CONFIGURATION BULLETINS — en-tête & modèle par établissement
# ════════════════════════════════════════════════════════════

def _require_admin_school(current_user: dict) -> int:
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Non autorisé")
    school_id = current_user.get("school_id")
    if not school_id:
        raise HTTPException(status_code=400, detail="Aucun établissement associé à ce compte")
    return school_id


@router.get("/bulletin-settings", response_model=BulletinSettingsResponse)
def get_bulletin_settings(
    current_user: dict = Depends(get_current_user),
    master_db: Session = Depends(get_db_session),
):
    from app.models.school import School
    from app.services.bulletin_templates import AVAILABLE_TEMPLATES, BULLETIN_SCOPES

    school_id = _require_admin_school(current_user)
    school = master_db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="Établissement non trouvé")

    templates = {k: v["label"] for k, v in AVAILABLE_TEMPLATES.items()}
    return BulletinSettingsResponse(
        school_id=school.id,
        name=school.name,
        logo_url=school.logo_url,
        bulletin_po_box=school.bulletin_po_box,
        bulletin_motto=school.bulletin_motto,
        bulletin_delegation_en=school.bulletin_delegation_en,
        bulletin_delegation_fr=school.bulletin_delegation_fr,
        bulletin_next_term_note=school.bulletin_next_term_note,
        bulletin_template=getattr(school, "bulletin_template", None) or "cameroon_bilingual",
        bulletin_scope=getattr(school, "bulletin_scope", None) or "trimestre",
        available_templates=templates,
        available_scopes=BULLETIN_SCOPES,
    )


@router.put("/bulletin-settings", response_model=BulletinSettingsResponse)
def update_bulletin_settings(
    data: BulletinSettingsUpdate,
    current_user: dict = Depends(get_current_user),
    master_db: Session = Depends(get_db_session),
):
    from app.models.school import School
    from app.services.bulletin_templates import AVAILABLE_TEMPLATES, BULLETIN_SCOPES

    school_id = _require_admin_school(current_user)
    school = master_db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="Établissement non trouvé")

    payload = data.model_dump(exclude_unset=True)
    if "bulletin_template" in payload and payload["bulletin_template"] not in AVAILABLE_TEMPLATES:
        raise HTTPException(status_code=400, detail="Modèle de bulletin invalide")
    if "bulletin_scope" in payload and payload["bulletin_scope"] not in BULLETIN_SCOPES:
        raise HTTPException(status_code=400, detail="Portée de bulletin invalide")

    for field, value in payload.items():
        setattr(school, field, value)
    school.updated_at = datetime.utcnow()
    master_db.commit()
    master_db.refresh(school)

    templates = {k: v["label"] for k, v in AVAILABLE_TEMPLATES.items()}
    return BulletinSettingsResponse(
        school_id=school.id,
        name=school.name,
        logo_url=school.logo_url,
        bulletin_po_box=school.bulletin_po_box,
        bulletin_motto=school.bulletin_motto,
        bulletin_delegation_en=school.bulletin_delegation_en,
        bulletin_delegation_fr=school.bulletin_delegation_fr,
        bulletin_next_term_note=school.bulletin_next_term_note,
        bulletin_template=getattr(school, "bulletin_template", None) or "cameroon_bilingual",
        bulletin_scope=getattr(school, "bulletin_scope", None) or "trimestre",
        available_templates=templates,
        available_scopes=BULLETIN_SCOPES,
    )


# ════════════════════════════════════════════════════════════
# STATS ADMIN — tableau de bord
# ════════════════════════════════════════════════════════════

@router.get("/stats")
def get_admin_stats(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
    master_db: Session = Depends(get_db_session),
):
    """Statistiques du tableau de bord administrateur"""
    from app.models.school import School, Professeur, Classe, Eleve, Matiere, Note, AnneeScolaire

    total_professeurs = db.query(Professeur).filter(Professeur.is_active == True).count()
    total_classes = db.query(Classe).count()
    total_eleves = db.query(Eleve).count()
    total_matieres = db.query(Matiere).count()
    total_notes = db.query(Note).count()
    annee_active = db.query(AnneeScolaire).filter(AnneeScolaire.is_active == True).first()

    school_name = None
    logo_url = None
    primary_color = "#10b981"
    secondary_color = "#f59e0b"
    school_id = current_user.get("school_id")
    if school_id:
        school = master_db.query(School).filter(School.id == school_id).first()
        if school:
            school_name = school.name
            logo_url = school.logo_url
            primary_color = school.primary_color or primary_color
            secondary_color = school.secondary_color or secondary_color

    return {
        "total_professeurs": total_professeurs,
        "total_classes": total_classes,
        "total_eleves": total_eleves,
        "total_matieres": total_matieres,
        "total_notes": total_notes,
        "annee_scolaire": annee_active.annee if annee_active else None,
        "school_name": school_name,
        "school_id": school_id,
        "logo_url": logo_url,
        "primary_color": primary_color,
        "secondary_color": secondary_color,
    }
