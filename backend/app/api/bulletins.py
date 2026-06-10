"""
API Bulletins scolaires — consultation, export et import par trimestre.
"""
import csv
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.auth.security import get_current_user
from app.db.connection import get_db_session
from app.db.multi_tenant import get_tenant_session
from app.models.school import (
    Eleve, Matiere, Note, AttributionProfesseur, Classe,
)
from app.services.bulletin_service import build_eleve_bulletin, build_classe_bulletins
from app.services.bulletin_cameroon import build_cameroon_bulletin
from app.services.bulletin_pdf_cameroon import build_cameroon_bulletin_pdf
from app.services.bulletin_templates import get_pdf_builder_name, resolve_template
from app.api.notes import (
    _find_note, _verifier_periode_saisie, _enforce_professor_deadline,
)

router = APIRouter(prefix="/bulletins", tags=["Bulletins"])


def _require_bulletin_access(current_user: dict):
    if current_user.get("role") not in ("admin", "professeur"):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et professeurs")


@router.get("/eleve/{eleve_id}")
def get_bulletin_eleve(
    eleve_id: int,
    trimestre: int = 1,
    format: str = "cameroon",
    lang: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
    master_db: Session = Depends(get_db_session),
):
    _require_bulletin_access(current_user)
    if trimestre not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="Trimestre invalide (1, 2 ou 3)")

    if format == "cameroon":
        data = build_cameroon_bulletin(
            db, eleve_id, trimestre,
            lang=lang,
            master_db=master_db,
            school_id=current_user.get("school_id"),
        )
    else:
        data = build_eleve_bulletin(db, eleve_id, trimestre)
    if "error" in data:
        raise HTTPException(status_code=404, detail=data["error"])
    return data


@router.get("/classe/{classe_id}")
def get_bulletins_classe(
    classe_id: int,
    trimestre: int = 1,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    _require_bulletin_access(current_user)
    if trimestre not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="Trimestre invalide (1, 2 ou 3)")

    data = build_classe_bulletins(db, classe_id, trimestre)
    if "error" in data:
        raise HTTPException(status_code=404, detail=data["error"])
    return data


def _bulletin_csv_rows(bulletin: dict) -> list:
    rows = []
    rows.append([
        "Élève", bulletin["eleve"], "Matricule", bulletin["matricule"],
        "Classe", bulletin.get("classe") or "", "Trimestre", bulletin["trimestre"],
        "Moyenne générale", bulletin["moyenne_generale"], "Mention", bulletin["mention"],
    ])
    rows.append([])
    rows.append([
        "Matière", "1ère séq.", "Coef.", "2ème séq.", "Coef.",
        "Note trim.", "Coef. trim.", "Moy. calculée", "Moy. matière",
    ])
    for d in bulletin["details_matieres"]:
        rows.append([
            d["matiere"],
            d["sequence_1"] if d["sequence_1"] is not None else "",
            d["coef_sequence_1"] if d["coef_sequence_1"] is not None else "",
            d["sequence_2"] if d["sequence_2"] is not None else "",
            d["coef_sequence_2"] if d["coef_sequence_2"] is not None else "",
            d["note_trimestre"] if d["note_trimestre"] is not None else "",
            d["coef_trimestre"] if d["coef_trimestre"] is not None else "",
            d["moyenne_calculee"] if d["moyenne_calculee"] is not None else "",
            d["moyenne_matiere"] if d["moyenne_matiere"] is not None else "",
        ])
    return rows


@router.get("/eleve/{eleve_id}/export/csv")
def export_bulletin_eleve_csv(
    eleve_id: int,
    trimestre: int = 1,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    _require_bulletin_access(current_user)
    bulletin = build_eleve_bulletin(db, eleve_id, trimestre)
    if "error" in bulletin:
        raise HTTPException(status_code=404, detail=bulletin["error"])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    for row in _bulletin_csv_rows(bulletin):
        writer.writerow(row)

    filename = f"bulletin_{bulletin['matricule']}_T{trimestre}.csv"
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/classe/{classe_id}/export/csv")
def export_bulletins_classe_csv(
    classe_id: int,
    trimestre: int = 1,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    _require_bulletin_access(current_user)
    data = build_classe_bulletins(db, classe_id, trimestre)
    if "error" in data:
        raise HTTPException(status_code=404, detail=data["error"])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Rang", "Nom", "Prénom", "Matricule", "Classe", "Trimestre",
        "Moyenne générale", "Mention",
    ])
    for b in data["bulletins"]:
        writer.writerow([
            b.get("rang", ""),
            b["eleve_nom"],
            b["eleve_prenom"],
            b["matricule"],
            b.get("classe") or data["classe"],
            trimestre,
            b["moyenne_generale"],
            b["mention"],
        ])
    writer.writerow([])
    writer.writerow(["Détail par matière"])
    writer.writerow([
        "Élève", "Matière", "1ère séq.", "Coef.", "2ème séq.", "Coef.",
        "Note trim.", "Coef. trim.", "Moy. matière",
    ])
    for b in data["bulletins"]:
        for d in b["details_matieres"]:
            writer.writerow([
                b["eleve"],
                d["matiere"],
                d["sequence_1"] if d["sequence_1"] is not None else "",
                d["coef_sequence_1"] if d["coef_sequence_1"] is not None else "",
                d["sequence_2"] if d["sequence_2"] is not None else "",
                d["coef_sequence_2"] if d["coef_sequence_2"] is not None else "",
                d["note_trimestre"] if d["note_trimestre"] is not None else "",
                d["coef_trimestre"] if d["coef_trimestre"] is not None else "",
                d["moyenne_matiere"] if d["moyenne_matiere"] is not None else "",
            ])

    filename = f"bulletins_{data['classe']}_T{trimestre}.csv"
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/classe/{classe_id}/export/xlsx")
def export_bulletins_classe_xlsx(
    classe_id: int,
    trimestre: int = 1,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Export Excel (.xlsx) — nécessite openpyxl."""
    _require_bulletin_access(current_user)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Export Excel indisponible (openpyxl non installé). Utilisez l'export CSV.",
        )

    data = build_classe_bulletins(db, classe_id, trimestre)
    if "error" in data:
        raise HTTPException(status_code=404, detail=data["error"])

    wb = Workbook()
    ws = wb.active
    ws.title = f"Bulletins T{trimestre}"

    header_font = Font(bold=True)
    ws.append([f"Bulletins — {data['classe']} — Trimestre {trimestre}"])
    ws.append([])
    ws.append(["Rang", "Nom", "Prénom", "Matricule", "Moyenne", "Mention"])
    for cell in ws[3]:
        cell.font = header_font

    for b in data["bulletins"]:
        ws.append([
            b.get("rang"),
            b["eleve_nom"],
            b["eleve_prenom"],
            b["matricule"],
            b["moyenne_generale"],
            b["mention"],
        ])

    ws2 = wb.create_sheet("Détail matières")
    ws2.append([
        "Élève", "Matière", "1ère séq.", "Coef.", "2ème séq.", "Coef.",
        "Note trim.", "Coef. trim.", "Moy. matière",
    ])
    for cell in ws2[1]:
        cell.font = header_font

    for b in data["bulletins"]:
        for d in b["details_matieres"]:
            ws2.append([
                b["eleve"],
                d["matiere"],
                d["sequence_1"],
                d["coef_sequence_1"],
                d["sequence_2"],
                d["coef_sequence_2"],
                d["note_trimestre"],
                d["coef_trimestre"],
                d["moyenne_matiere"],
            ])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bulletins_{data['classe']}_T{trimestre}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_bulletin_pdf(bulletin: dict) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise HTTPException(
            status_code=501,
            detail="Export PDF indisponible (reportlab non installé).",
        ) from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, spaceAfter=8)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.grey)
    elements = []

    elements.append(Paragraph("BULLETIN SCOLAIRE", title_style))
    elements.append(Paragraph(
        f"<b>{bulletin['eleve']}</b> — Matricule : {bulletin.get('matricule', '—')}",
        styles["Normal"],
    ))
    meta_parts = [
        f"Classe : {bulletin.get('classe') or '—'}",
        f"Trimestre : {bulletin['trimestre']}",
    ]
    if bulletin.get("annee_scolaire"):
        meta_parts.append(f"Année : {bulletin['annee_scolaire']}")
    if bulletin.get("rang"):
        meta_parts.append(f"Rang : {bulletin['rang']}")
    elements.append(Paragraph(" | ".join(meta_parts), sub_style))
    elements.append(Spacer(1, 0.4 * cm))

    moyenne = bulletin["moyenne_generale"]
    mention = bulletin.get("mention", "")
    elements.append(Paragraph(
        f"<b>Moyenne générale : {moyenne} / 20</b> — {mention}",
        styles["Heading2"],
    ))
    elements.append(Spacer(1, 0.5 * cm))

    table_data = [[
        "Matière", "1ère séq.", "Coef.", "2ème séq.", "Coef.",
        "Note trim.", "Coef.", "Moy. matière",
    ]]
    for d in bulletin.get("details_matieres", []):
        table_data.append([
            d["matiere"],
            str(d["sequence_1"]) if d["sequence_1"] is not None else "—",
            str(d["coef_sequence_1"]) if d["coef_sequence_1"] is not None else "—",
            str(d["sequence_2"]) if d["sequence_2"] is not None else "—",
            str(d["coef_sequence_2"]) if d["coef_sequence_2"] is not None else "—",
            str(d["note_trimestre"]) if d["note_trimestre"] is not None else "—",
            str(d["coef_trimestre"]) if d["coef_trimestre"] is not None else "—",
            str(d["moyenne_matiere"]) if d["moyenne_matiere"] is not None else "—",
        ])

    if len(table_data) == 1:
        table_data.append(["Aucune note pour ce trimestre", "", "", "", "", "", "", ""])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.8 * cm))
    elements.append(Paragraph("Document généré par EduSaaS — Bloomarone", sub_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


TYPE_ALIASES = {
    "sequence_1": "sequence_1",
    "seq1": "sequence_1",
    "seq 1": "sequence_1",
    "1ere sequence": "sequence_1",
    "1ère séquence": "sequence_1",
    "1ere séquence": "sequence_1",
    "premiere sequence": "sequence_1",
    "sequence_2": "sequence_2",
    "seq2": "sequence_2",
    "seq 2": "sequence_2",
    "2eme sequence": "sequence_2",
    "2ème séquence": "sequence_2",
    "2eme séquence": "sequence_2",
    "deuxieme sequence": "sequence_2",
    "trimestre": "trimestre",
    "note trimestrielle": "trimestre",
    "note trim": "trimestre",
}


def _normalize_type(raw: str) -> Optional[str]:
    if not raw:
        return None
    key = str(raw).strip().lower()
    return TYPE_ALIASES.get(key)


def _parse_header_row(row: list) -> dict:
    """Mappe les colonnes par libellé (insensible à la casse)."""
    mapping = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        label = str(cell).strip().lower()
        if label in ("matricule", "matricule élève", "matricule eleve"):
            mapping["matricule"] = idx
        elif label in ("matière", "matiere", "matiere ", "subject"):
            mapping["matiere"] = idx
        elif label in ("type", "type évaluation", "type evaluation", "evaluation"):
            mapping["type"] = idx
        elif label in ("note", "valeur", "score"):
            mapping["note"] = idx
        elif label in ("coefficient", "coef", "coef."):
            mapping["coef"] = idx
    return mapping


@router.get("/import/template.xlsx")
def download_import_template(
    current_user: dict = Depends(get_current_user),
):
    """Télécharge un modèle Excel pour l'import des notes (bulletins)."""
    _require_bulletin_access(current_user)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Modèle Excel indisponible (openpyxl non installé).",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Import notes"
    headers = ["Matricule", "Matière", "Type", "Note", "Coefficient"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["ELEVE001", "Mathématiques", "sequence_1", 14.5, 1])
    ws.append(["ELEVE001", "Mathématiques", "sequence_2", 12, 1])
    ws.append(["ELEVE001", "Français", "trimestre", 15, 2])
    ws.append([])
    ws.append(["Types acceptés : sequence_1, sequence_2, trimestre"])
    ws.append(["Note entre 0 et 20 — Trimestre choisi à l'import"])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modele_import_bulletins.xlsx"'},
    )


@router.post("/import/xlsx")
async def import_bulletins_xlsx(
    classe_id: int,
    trimestre: int = 1,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
):
    """Importe des notes depuis Excel pour alimenter les bulletins."""
    _require_bulletin_access(current_user)
    if trimestre not in (1, 2, 3):
        raise HTTPException(status_code=400, detail="Trimestre invalide (1, 2 ou 3)")

    classe = db.query(Classe).filter(Classe.id == classe_id).first()
    if not classe:
        raise HTTPException(status_code=404, detail="Classe non trouvée")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Fichier Excel (.xlsx) requis")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=501, detail="Import Excel indisponible (openpyxl non installé).")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Fichier Excel invalide : {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Fichier vide")

    header_idx = None
    col_map = {}
    for i, row in enumerate(rows[:10]):
        candidate = _parse_header_row(list(row))
        if "matricule" in candidate and "matiere" in candidate and "note" in candidate:
            header_idx = i
            col_map = candidate
            break

    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="En-têtes requis : Matricule, Matière, Type, Note (Coefficient optionnel)",
        )

    role = current_user.get("role")
    professeur_id = current_user.get("id")

    eleves = {
        e.matricule.strip().upper(): e
        for e in db.query(Eleve).filter(Eleve.classe_id == classe_id).all()
    }
    matieres = {
        m.nom.strip().lower(): m
        for m in db.query(Matiere).all()
    }

    imported = 0
    updated = 0
    errors = []

    for line_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            matricule = str(row[col_map["matricule"]]).strip().upper()
            matiere_nom = str(row[col_map["matiere"]]).strip().lower()
            note_val = row[col_map["note"]]
            type_raw = row[col_map["type"]] if "type" in col_map else "sequence_1"
            coef_raw = row[col_map["coef"]] if "coef" in col_map else 1.0

            if not matricule or matricule.startswith("TYPES"):
                continue

            eleve = eleves.get(matricule)
            if not eleve:
                errors.append(f"Ligne {line_no} : matricule « {matricule} » introuvable dans la classe")
                continue

            matiere = matieres.get(matiere_nom)
            if not matiere:
                errors.append(f"Ligne {line_no} : matière « {row[col_map['matiere']]} » introuvable")
                continue

            type_eval = _normalize_type(str(type_raw)) if type_raw else "sequence_1"
            if not type_eval:
                errors.append(f"Ligne {line_no} : type « {type_raw} » invalide")
                continue

            try:
                valeur = float(note_val)
            except (TypeError, ValueError):
                errors.append(f"Ligne {line_no} : note invalide « {note_val} »")
                continue

            if valeur < 0 or valeur > 20:
                errors.append(f"Ligne {line_no} : note hors plage (0-20)")
                continue

            try:
                coefficient = float(coef_raw) if coef_raw is not None else 1.0
            except (TypeError, ValueError):
                coefficient = 1.0

            attribution = db.query(AttributionProfesseur).filter(
                and_(
                    AttributionProfesseur.classe_id == classe_id,
                    AttributionProfesseur.matiere_id == matiere.id,
                    AttributionProfesseur.is_active == True,
                )
            ).first()
            if not attribution:
                errors.append(f"Ligne {line_no} : pas d'attribution pour {matiere.nom}")
                continue

            if role == "professeur":
                prof_attr = db.query(AttributionProfesseur).filter(
                    and_(
                        AttributionProfesseur.professeur_id == professeur_id,
                        AttributionProfesseur.classe_id == classe_id,
                        AttributionProfesseur.matiere_id == matiere.id,
                        AttributionProfesseur.is_active == True,
                    )
                ).first()
                if not prof_attr:
                    errors.append(f"Ligne {line_no} : matière non assignée à ce professeur")
                    continue
                verification = _verifier_periode_saisie(db, classe_id, matiere.id)
                try:
                    _enforce_professor_deadline(role, verification)
                except HTTPException:
                    errors.append(f"Ligne {line_no} : période de saisie fermée pour {matiere.nom}")
                    continue

            effective_prof = professeur_id if role == "professeur" else attribution.professeur_id
            existing = _find_note(
                db, eleve.id, matiere.id, trimestre, type_eval, professeur_id=effective_prof,
            )

            if existing:
                existing.valeur = valeur
                existing.coefficient = coefficient
                existing.date_saisie = datetime.utcnow()
                updated += 1
            else:
                db.add(Note(
                    eleve_id=eleve.id,
                    matiere_id=matiere.id,
                    professeur_id=effective_prof,
                    trimestre=trimestre,
                    type_evaluation=type_eval,
                    valeur=valeur,
                    coefficient=coefficient,
                    date_creation=datetime.utcnow(),
                    date_saisie=datetime.utcnow(),
                ))
                imported += 1

        except Exception as exc:
            errors.append(f"Ligne {line_no} : {exc}")

    db.commit()

    return {
        "success": True,
        "classe": classe.nom,
        "trimestre": trimestre,
        "imported": imported,
        "updated": updated,
        "total": imported + updated,
        "errors": errors[:50],
        "error_count": len(errors),
    }


@router.get("/eleve/{eleve_id}/export/pdf")
def export_bulletin_eleve_pdf(
    eleve_id: int,
    trimestre: int = 1,
    template: str = "cameroon",
    lang: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_tenant_session),
    master_db: Session = Depends(get_db_session),
):
    _require_bulletin_access(current_user)
    school_id = current_user.get("school_id")
    from app.models.school import School, Classe, Eleve

    resolved = template
    if template == "auto" and school_id:
        school_row = master_db.query(School).filter(School.id == school_id).first()
        eleve = db.query(Eleve).filter(Eleve.id == eleve_id).first()
        section = None
        if eleve and eleve.classe_id:
            classe = db.query(Classe).filter(Classe.id == eleve.classe_id).first()
            section = getattr(classe, "section", None) if classe else None
        resolved = resolve_template(
            getattr(school_row, "bulletin_template", None) if school_row else None,
            section,
        )
    elif template == "cameroon":
        resolved = "cameroon_bilingual"

    builder = get_pdf_builder_name(resolved)
    if builder == "cameroon":
        bulletin = build_cameroon_bulletin(
            db, eleve_id, trimestre,
            lang=lang,
            master_db=master_db,
            school_id=school_id,
        )
        if "error" in bulletin:
            raise HTTPException(status_code=404, detail=bulletin["error"])
        try:
            pdf_bytes = build_cameroon_bulletin_pdf(bulletin)
        except ImportError as exc:
            raise HTTPException(status_code=501, detail="Export PDF indisponible (reportlab).") from exc
    else:
        bulletin = build_eleve_bulletin(db, eleve_id, trimestre)
        if "error" in bulletin:
            raise HTTPException(status_code=404, detail=bulletin["error"])
        pdf_bytes = _build_bulletin_pdf(bulletin)

    suffix = bulletin.get("lang", "fr") if builder == "cameroon" else "std"
    filename = f"bulletin_{bulletin['matricule']}_T{trimestre}_{suffix}.pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
