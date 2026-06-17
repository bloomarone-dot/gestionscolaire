"""Import / export de listes d'élèves (Excel ou CSV)."""
from __future__ import annotations

import csv
import io
import re
from typing import BinaryIO, Optional

from sqlalchemy.orm import Session

from app import crud
from app.models import Eleve, Parent
from app.schemas import EleveCreate, EleveUpdate, ParentIn

HEADER_ALIASES = {
    "matricule": {"matricule", "mat", "code", "id"},
    "nom": {"nom", "name", "nom_famille", "nom de famille"},
    "prenom": {"prenom", "prénom", "firstname", "first name"},
    "sexe": {"sexe", "genre", "sex", "m/f"},
    "classe": {"classe", "class", "classe_id", "class_id", "classe id"},
    "parent_nom": {"parent", "parent_nom", "parent nom", "tuteur", "nom parent", "nom_parent"},
    "parent_phone": {"telephone", "téléphone", "phone", "tel", "parent_phone", "parent telephone", "parent tel", "contact"},
    "parent_phone2": {"telephone2", "téléphone2", "phone2", "tel2", "parent_phone2", "parent telephone 2"},
    "parent_adresse": {"adresse", "address", "parent_adresse"},
}


def _norm_header(value) -> str:
    s = str(value or "").strip().lower()
    s = s.replace("é", "e").replace("è", "e").replace("ê", "e")
    return re.sub(r"\s+", " ", s)


def _parse_header_row(row: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = _norm_header(cell)
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _cell(row: list, col_map: dict[str, int], field: str):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_sexe(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    v = value.strip().upper()
    if v in ("M", "MASCULIN", "H", "GARCON", "GARÇON"):
        return "M"
    if v in ("F", "FEMININ", "FÉMININ", "FILLE"):
        return "F"
    return None


def _read_rows_xlsx(content: bytes) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _read_rows_csv(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def read_tabular_rows(content: bytes, filename: str) -> list[list]:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xls")):
        return _read_rows_xlsx(content)
    return _read_rows_csv(content)


def build_template_xlsx(
    *,
    classe_nom: Optional[str] = None,
    section: Optional[str] = None,
) -> bytes:
    """Modèle Excel — sans colonne Classe si import ciblé sur une classe."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Eleves"
    per_class = bool(classe_nom)
    headers = [
        "Matricule",
        "Nom",
        "Prenom",
        "Sexe",
        *([] if per_class else ["Classe"]),
        "Parent nom",
        "Parent telephone",
        "Parent telephone 2",
        "Parent adresse",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    sample_class = classe_nom or "Tle D1"
    if per_class:
        ws.append(["", "Nkomo", "Jean", "M", "Nkomo Paul", "699112233", "", "Yaounde"])
        ws.append(["", "Fotso", "Marie", "F", "Fotso Anne", "677445566", "", ""])
    else:
        ws.append(["", "Nkomo", "Jean", "M", sample_class, "Nkomo Paul", "699112233", "", "Yaounde"])
        ws.append(["", "Fotso", "Marie", "F", sample_class, "Fotso Anne", "677445566", "", ""])
    ws.append([])
    ws.append(["Matricule : laisser vide pour generer automatiquement"])
    if per_class:
        section_hint = f" ({section})" if section else ""
        ws.append([f"Classe cible : {classe_nom}{section_hint} — tous les eleves iront dans cette classe"])
    else:
        ws.append(["Classe : nom personnalise (ex. Tle D1) ou ID numerique"])
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 40)
        ws.column_dimensions[col[0].column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rows(eleves: list[Eleve], classe_names: dict[int, str]) -> list[dict]:
    rows = []
    for e in eleves:
        parent = e.parents[0] if e.parents else None
        rows.append({
            "matricule": e.matricule,
            "nom": e.nom,
            "prenom": e.prenom or "",
            "sexe": e.sexe or "",
            "classe": classe_names.get(e.classe_id, str(e.classe_id or "")),
            "parent_nom": parent.nom if parent else "",
            "parent_phone": parent.phone if parent else "",
            "parent_phone2": parent.phone2 if parent else "",
            "parent_adresse": parent.adresse if parent else "",
        })
    return rows


def rows_to_csv(rows: list[dict]) -> bytes:
    fields = [
        "matricule", "nom", "prenom", "sexe", "classe",
        "parent_nom", "parent_phone", "parent_phone2", "parent_adresse",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    for row in rows:
        writer.writerow({f: row.get(f, "") for f in fields})
    return buf.getvalue().encode("utf-8-sig")


def rows_to_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Eleves"
    headers = [
        "Matricule", "Nom", "Prenom", "Sexe", "Classe",
        "Parent nom", "Parent telephone", "Parent telephone 2", "Parent adresse",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([
            row.get("matricule", ""),
            row.get("nom", ""),
            row.get("prenom", ""),
            row.get("sexe", ""),
            row.get("classe", ""),
            row.get("parent_nom", ""),
            row.get("parent_phone", ""),
            row.get("parent_phone2", ""),
            row.get("parent_adresse", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_classe_id(
    raw: Optional[str],
    default_classe_id: Optional[int],
    classe_lookup: dict[str, int],
) -> Optional[int]:
    if raw:
        if raw.isdigit():
            return int(raw)
        key = raw.strip().lower()
        if key in classe_lookup:
            return classe_lookup[key]
        return None
    return default_classe_id


def import_rows(
    db: Session,
    tenant_id: int,
    rows: list[list],
    *,
    default_classe_id: Optional[int] = None,
    force_classe_id: Optional[int] = None,
    classe_lookup: Optional[dict[str, int]] = None,
) -> dict:
    """Importe ou met à jour des élèves par matricule.

    Si ``force_classe_id`` est défini (import par classe), tous les élèves
    sont inscrits dans cette classe ; la colonne Classe du fichier est ignorée.
    """
    lookup = classe_lookup or {}
    target_classe_id = force_classe_id if force_classe_id is not None else default_classe_id
    if not rows:
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": ["Fichier vide"]}

    header_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows[:15]):
        candidate = _parse_header_row(list(row))
        if "nom" in candidate:
            header_idx = i
            col_map = candidate
            break

    if header_idx is None:
        return {
            "imported": 0, "updated": 0, "skipped": 0,
            "errors": ["En-têtes requis : au minimum une colonne « Nom »"],
        }

    imported = updated = skipped = 0
    errors: list[str] = []

    for line_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        nom = _cell(row, col_map, "nom")
        if not nom:
            hint = _cell(row, col_map, "matricule") or ""
            if hint.lower().startswith("matricule") or hint.lower().startswith("classe"):
                continue
            errors.append(f"Ligne {line_no} : nom obligatoire")
            skipped += 1
            continue

        matricule = _cell(row, col_map, "matricule")
        prenom = _cell(row, col_map, "prenom")
        sexe = _normalize_sexe(_cell(row, col_map, "sexe"))
        classe_raw = _cell(row, col_map, "classe")
        if force_classe_id is not None:
            classe_id = force_classe_id
            if classe_raw:
                resolved = _resolve_classe_id(classe_raw, None, lookup)
                if resolved and resolved != force_classe_id:
                    errors.append(
                        f"Ligne {line_no} : classe « {classe_raw} » ignorée "
                        f"(import pour la classe sélectionnée)"
                    )
        else:
            classe_id = _resolve_classe_id(classe_raw, default_classe_id, lookup)

        if classe_id is None:
            errors.append(
                f"Ligne {line_no} ({nom}) : classe obligatoire — "
                "sélectionnez la classe à l'import ou renseignez la colonne Classe"
            )
            skipped += 1
            continue

        parent_nom = _cell(row, col_map, "parent_nom")
        parent_phone = _cell(row, col_map, "parent_phone")
        parent_phone2 = _cell(row, col_map, "parent_phone2")
        parent_adresse = _cell(row, col_map, "parent_adresse")

        existing = None
        if matricule:
            existing = crud.get_eleve_by_matricule(db, tenant_id, matricule)

        try:
            if existing:
                crud.update_eleve(db, tenant_id, existing.id, EleveUpdate(
                    nom=nom,
                    prenom=prenom,
                    sexe=sexe,
                    classe_id=classe_id,
                ))
                if parent_nom and parent_phone:
                    _upsert_parent(db, tenant_id, existing.id, parent_nom, parent_phone, parent_phone2, parent_adresse)
                updated += 1
            else:
                parents = []
                if parent_nom and parent_phone:
                    parents.append(ParentIn(
                        nom=parent_nom,
                        phone=parent_phone,
                        phone2=parent_phone2,
                        adresse=parent_adresse,
                    ))
                crud.create_eleve(db, tenant_id, EleveCreate(
                    nom=nom,
                    prenom=prenom,
                    matricule=matricule or None,
                    sexe=sexe,
                    classe_id=classe_id,
                    parents=parents,
                ))
                imported += 1
        except Exception as exc:
            errors.append(f"Ligne {line_no} ({nom}) : {exc}")
            skipped += 1

    return {"imported": imported, "updated": updated, "skipped": skipped, "errors": errors}


def _upsert_parent(
    db: Session,
    tenant_id: int,
    eleve_id: int,
    nom: str,
    phone: str,
    phone2: Optional[str],
    adresse: Optional[str],
) -> None:
    eleve = crud.get_eleve(db, tenant_id, eleve_id)
    if eleve.parents:
        p = eleve.parents[0]
        p.nom = nom
        p.phone = phone
        p.phone2 = phone2
        p.adresse = adresse
    else:
        db.add(Parent(
            tenant_id=tenant_id,
            eleve_id=eleve_id,
            nom=nom,
            phone=phone,
            phone2=phone2,
            adresse=adresse,
        ))
    db.commit()
